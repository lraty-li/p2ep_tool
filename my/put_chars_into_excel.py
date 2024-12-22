import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from common import getCharsFolderOfPgm, pgmTargets, workPalceFolder, pgmSzie
from openpyxl.styles import Alignment

# 创建一个新的工作簿
workBook = Workbook()
sheet = workBook.active
workBook.remove(sheet)

imgShowSizeBase = 12
imgShowSizeFactor = 10
imgShowSize = imgShowSizeBase * imgShowSizeFactor
gridSize = imgShowSize / 7.5
centerAlign = Alignment(horizontal="center", vertical="center")

for pgmIndex, pgm in enumerate(pgmTargets, start=1):
    workSheet = workBook.create_sheet(pgm, pgmIndex)
    pgmCharsFolder = os.path.join(workPalceFolder, getCharsFolderOfPgm(pgm))

    # 获取文件夹中所有图片文件
    image_files = os.listdir(pgmCharsFolder)
    workSheet.column_dimensions["A"].width = gridSize
    pgmSzieConfig = pgmSzie[pgmIndex - 1]
    for imgIndex in range(len(image_files)):
        image_path = os.path.join(pgmCharsFolder, "{}.png".format(imgIndex))
        # 创建一个 Image 对象
        img = Image(image_path)

        # 设置图片大小 (可选，根据需要调整)
        img.width, img.height = (
            img.width * imgShowSizeFactor,
            img.height * imgShowSizeFactor,
        )

        # 插入图片到对应单元格
        cellIndex = imgIndex + 1
        cellAddress = f"A{cellIndex}"
        workSheet.add_image(img, cellAddress)
        cell = workSheet.cell(cellIndex, 1)
        cellForText = workSheet.cell(cellIndex, 2)
        cell.alignment = centerAlign
        cellForText.alignment = centerAlign

        # 我也不知道excel宽高是为什么不同单位
        workSheet.row_dimensions[cellIndex].height = imgShowSize


workBook.save("persona2_ep_charset.xlsx")
